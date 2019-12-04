#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@File : excel.py
@Contact : yangxx@civaonline.cn
@MTime : 2019/7/3 11:19
@Author: yangxingxiang
@Version: 1.0
@Description: 封装一些常用的excel操作方法
"""
import os
import time
from openpyxl import Workbook
# from openpyxl.styles import colors
# from openpyxl.styles import PatternFill
from common.helper import make_dir

t = time.strftime('%Y-%m-%d_%H-%M-%S')


class Excel(object):
    def __init__(self, excel_name, parent_folder=None):
        """
        :param excel_name: excel文件名
        :param parent_folder: 文件夹
        """
        self.excel_name = '{}_{}.xlsx'.format(excel_name, t)
        self.parent_folder = parent_folder
        self.work_book = Workbook()
        make_dir(self.parent_folder)

    @property
    def file_path(self):
        file_path = os.path.join(self.parent_folder, self.excel_name)
        return file_path

    def create_sheet(self, sheet_name, index: int = None):
        try:
            self.work_book.create_sheet(title=sheet_name, index=index)
        except Exception as err:
            print('创建Sheet错误:', err)

    def rename_sheet(self, old_name, new_name):
        try:
            sheet_to_rename = self.work_book[old_name]
            sheet_to_rename.title = new_name
        except Exception as err:
            print('更改Sheet名称错误:', err)

    def copy_sheet(self, sheet_name):
        try:
            source = self.work_book[sheet_name]
            self.work_book.copy_worksheet(from_worksheet=source)
        except Exception as err:
            print('复制Sheet错误:', err)

    # def set_font_color(self):
    #     fill = PatternFill(bgColor=colors.GREEN)
    #     return fill

    def get_sheet_row_count(self, sheet_name):
        try:
            # self.work_book.get_active_sheet()
            sheet = self.work_book.get_sheet_by_name(sheet_name)
            row_count = sheet.max_row
            return row_count
        except Exception as err:
            print('获取Sheet行数错误:', err)

    def get_sheet_column_count(self, sheet_name):
        try:
            # self.work_book.get_active_sheet()
            sheet = self.work_book.get_sheet_by_name(sheet_name)
            column_count = sheet.max_column
            return column_count
        except Exception as err:
            print('获取Sheet列数错误:', err)

    def write_rows(self, sheet_name, data: list, append=True):
        '''
        按行存入一批数据
        :param sheet_name: 表单名
        :param data: 二维数组，第一维是行
        :param append: 是否追加，默认为是，即从已有数据行往后追加写入，如果为否，则从第一行开始写入
        :return: None，如果写入失败会报错
        '''
        sheet = self.work_book.get_sheet_by_name(sheet_name)
        start_row = 1
        if append:
            start_row = sheet.max_row + 1
        for value in data:
            self.write_row(sheet_name, value, start_row)
            start_row += 1
        self.work_book.save(self.file_path)

    def write_row(self, sheet_name, data: list, row_index=1):
        sheet = self.work_book.get_sheet_by_name(sheet_name)
        for i, value in enumerate(data):
            # 从第一列开始写，直到将数据写完
            sheet.cell(row_index, i + 1, value=value)

    def write_columns(self, sheet_name, data: list, append=True):
        '''
        按列存入一批数据
        :param sheet_name: 表单名
        :param data: 二维数组，第一维是列
        :param append: 是否追加，默认为是，即从已有数据行往后追加写入，如果为否，则从第一行开始写入
        :return: None，如果写入失败会报错
        '''
        sheet = self.work_book.get_sheet_by_name(sheet_name)
        start_col = 1
        if append:
            start_col = sheet.max_column + 1
        for value in data:
            self.write_column(sheet_name, value, start_col)
            start_col += 1
        self.work_book.save(self.file_path)

    def write_column(self, sheet_name, data: list, col_index=1):
        sheet = self.work_book.get_sheet_by_name(sheet_name)
        for i, value in enumerate(data):
            # 从第一行开始写，直到将数据写完
            sheet.cell(i + 1, col_index, value=value)

    def save(self):
        self.work_book.save(self.file_path)
